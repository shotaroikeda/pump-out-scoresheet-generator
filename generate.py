from parse_pump_out import read_database
from parse_config import parse_config, titles_to_ids, config_all

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule, FormulaRule, Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter as gcl

import datetime
import sys

###### PUT THE PATH OF YOUR DATABASE HERE #####
DBPATH = 'pumpout-2020-05-03-22-18-1588558611155.db'
###### PUT THE PATH OF YOUR DATABASE HERE #####

def adjust_column_widths(ws, cols, rows):
	for c in cols:
		width = 0
		for r in rows:
			val = ws.cell(row=r, column=c).value
			if val == None: continue
			width = max(width, len(str(val)))
		ws.column_dimensions[gcl(c)].width = width + 1

def write_data_sheet(ws, db, chart_set, config):
	if chart_set == None:
		chart_set = set(db.charts)
	if config == None:
		config = config_all(db)

	latest_filtered_mix = -1
	for mid in config.mix_ids:
		if latest_filtered_mix == -1:
			latest_filtered_mix = mid
		elif db.mixes[mid].order > db.mixes[latest_filtered_mix].order:
			latest_filtered_mix = mid
	fver = db.newest_version_from_mix(latest_filtered_mix)

	charts = list(chart_set)
	charts.sort(key=lambda cid: db.chart_sort_key(cid, fver, down=True))

	headers = [
		"CID", # A
		"SID", # B
		"GID", # C
		"Title", # D
		"Cut", # E
		"Mode", # F
		"Difficulty", # G
		"First Seen", # H
		"Last Seen", # I
		"BPM", # J
		"Category", # K
		"Stepmaker", # L
	]
	headers += config.mixes
	headers += [
		"Labels",
		"Card",
		"Comment"
	]
	m = len(config.mixes)
	for i in range(len(headers)):
		ws.cell(row=1, column=i+1, value=headers[i]).font = Font(bold=True)
	for i, cid in enumerate(charts):
		sid = db.charts[cid].songId

		game_id = db.song_game_id(sid, fver)
		if game_id == None: game_id = ""

		first_seen = last_seen = "???"
		vid = db.chart_introduced(cid)
		if vid != None:
			first_seen = db.version_title(vid)
		vid = db.chart_last_seen(cid)
		if vid != None:
			last_seen = db.version_title(vid)

		ws.cell(row=i+2, column=1, value=cid)
		ws.cell(row=i+2, column=2, value=sid)
		ws.cell(row=i+2, column=3, value=game_id)
		ws.cell(row=i+2, column=4, value=db.song_title(sid, fver))
		ws.cell(row=i+2, column=5, value=db.song_cut_str(sid))
		ws.cell(row=i+2, column=6, value=db.chart_mode_str(cid, fver))
		ws.cell(row=i+2, column=7, value=db.chart_difficulty_str(cid, fver))
		ws.cell(row=i+2, column=8, value=first_seen)
		ws.cell(row=i+2, column=9, value=last_seen)
		ws.cell(row=i+2, column=10, value=db.song_bpm_str(sid, fver))
		ws.cell(row=i+2, column=11, value=db.song_category(sid, fver))
		ws.cell(row=i+2, column=12, value=str(db.chart_stepmaker(cid)))
		for j, mid in enumerate(config.mix_ids):
			ws.cell(row=i+2, column=13+j, value="NY"[db.chart_in_mix(cid, mid)])
		ws.cell(row=i+2, column=13+m, value=",".join(db.chart_labels(cid, fver)))
		ws.cell(row=i+2, column=14+m, value=db.song_card(sid, fver))
		ws.cell(row=i+2, column=15+m, value=db.song_comment(sid, fver))

	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 5
	ws.column_dimensions['C'].width = 5
	ws.column_dimensions['D'].width = 36
	ws.column_dimensions['E'].width = 9
	ws.column_dimensions['F'].width = 9
	ws.column_dimensions['G'].width = 3
	ws.column_dimensions['H'].width = 17
	ws.column_dimensions['I'].width = 17
	ws.column_dimensions['J'].width = 8
	ws.column_dimensions['K'].width = 14
	ws.column_dimensions['L'].width = 31
	for i in range(m):
		ws.column_dimensions[gcl(13+i)].width = 2
	ws.column_dimensions[gcl(13+m)].width = 39
	ws.column_dimensions[gcl(14+m)].width = 48
	ws.column_dimensions[gcl(15+m)].width = 120

	ws.freeze_panes = 'A2'

def write_score_sheet(ws, db, chart_set, config, mixId):
	latest_filtered_mix = -1
	for mid in config.mix_ids:
		if latest_filtered_mix == -1:
			latest_filtered_mix = mid
		elif db.mixes[mid].order > db.mixes[latest_filtered_mix].order:
			latest_filtered_mix = mid
	fver = db.newest_version_from_mix(latest_filtered_mix)

	charts = list(chart_set)
	charts.sort(key=lambda cid: db.chart_sort_key(cid, fver, down=True))

	mixes = config.mix_ids
	if len(mixes) == 1:
		mixes = []

	headers = [
		"CID",           # A
		"Title",         # B
		"Cut",           # C
		"Mode",          # D
		"Difficulty",    # E
	]
	if config.pad:
		headers += [
			"Passed (Pad)",  # F
			"Grade (Pad)",   # G
			"Miss (Pad)",    # H
			"Comment (Pad)", # I
		]
	if config.keyboard:
		headers += [
			"Passed (Kbd)",  # F J
			"Grade (Kbd)",   # G K
			"Miss (Kbd)",    # H L
			"Comment (Kbd)"  # I M
		]
	headers += [db.mixes[m].title for m in mixes] # F J N
	bold = Font(bold=True)
	gray = PatternFill("solid", fgColor="EEEEEE")
	dgray = PatternFill("solid", fgColor="CCCCCC")

	mix_col = 6
	border_cols = [5]
	if config.pad or config.keyboard:
		mix_col = 10
		border_cols += [9]
	if config.pad and config.keyboard:
		mix_col = 14
		border_cols += [13]

	for i in range(len(headers)):
		right = None
		c = ws.cell(row=1, column=i+1, value=headers[i])
		c.font = bold
		c.fill = dgray
		c.border = Border(bottom=Side(style="thick"), right=right)
	for i, cid in enumerate(charts):
		ws.cell(row=i+2, column=1, value=cid).fill = dgray
		ws.cell(row=i+2, column=2, value="=VLOOKUP(A%d, 'Data (Complete)'!A1:O9999, 4, FALSE)" % (i+2)).fill = gray
		ws.cell(row=i+2, column=3, value="=VLOOKUP(A%d, 'Data (Complete)'!A1:O9999, 5, FALSE)" % (i+2)).fill = gray
		ws.cell(row=i+2, column=4, value="=VLOOKUP(A%d, 'Data (Complete)'!A1:O9999, 6, FALSE)" % (i+2)).fill = gray
		ws.cell(row=i+2, column=5, value="=VLOOKUP(A%d, 'Data (Complete)'!A1:O9999, 7, FALSE)" % (i+2)).fill = gray
		for j, mid in enumerate(mixes):
			ws.cell(row=i+2, column=mix_col+j, value="NY"[db.chart_in_mix(cid, mid)]).fill = gray

	# TODO: This is not the horizontal style I want
	for i in range(len(charts)):
		ws.cell(row=i+2, column=9).alignment = Alignment(horizontal='fill')

	for c in range(len(headers)):
		for r in range(len(charts)):
			right = None
			if c+1 == len(headers):
				right = Side(style="thin")
			if c+1 in border_cols:
				right = Side(style="thick")
			border = Border(bottom=Side(style="thin", color="888888"), right=right)
			ws.cell(row=r+2, column=c+1).border = border

	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 30
	ws.column_dimensions['C'].width = 9
	ws.column_dimensions['D'].width = 9
	ws.column_dimensions['E'].width = 3
	c = 6
	if config.pad:
		ws.column_dimensions[gcl(c+0)].width = 4
		ws.column_dimensions[gcl(c+1)].width = 4
		ws.column_dimensions[gcl(c+2)].width = 4
		ws.column_dimensions[gcl(c+3)].width = 20
		c += 4
	if config.keyboard:
		ws.column_dimensions[gcl(c+0)].width = 4
		ws.column_dimensions[gcl(c+1)].width = 4
		ws.column_dimensions[gcl(c+2)].width = 4
		ws.column_dimensions[gcl(c+3)].width = 20
		c += 4
	for i in range(len(mixes)):
		ws.column_dimensions[gcl(c+i)].width = 2

	green = PatternFill("solid", bgColor="44FF44")
	red = PatternFill("solid", bgColor="FF4444")
	grade_f = PatternFill("solid", bgColor="555555")
	grade_d = PatternFill("solid", bgColor="666666")
	grade_c = PatternFill("solid", bgColor="777777")
	grade_b = PatternFill("solid", bgColor="888888")
	grade_a = PatternFill("solid", bgColor="AAAAAA")
	grade_s = PatternFill("solid", bgColor="DD88FF")
	grade_ss = PatternFill("solid", bgColor="FFEE00")
	grade_sss = PatternFill("solid", bgColor="44FF44")

	if config.pad or config.keyboard:
		ws.conditional_formatting.add('F2:F9999', CellIsRule(operator='equal', formula=['"Y"'], fill=green))
		ws.conditional_formatting.add('F2:F9999', CellIsRule(operator='equal', formula=['"N"'], fill=red))

	if config.pad and config.keyboard:
		ws.conditional_formatting.add('J2:J9999', CellIsRule(operator='equal', formula=['"Y"'], fill=green))
		ws.conditional_formatting.add('J2:J9999', CellIsRule(operator='equal', formula=['"N"'], fill=red))

	if config.pad or config.keyboard:
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"SSS"'], fill=grade_sss))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"SS"'], fill=grade_ss))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"S"'], fill=grade_s))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"A"'], fill=grade_a))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"B"'], fill=grade_b))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"C"'], fill=grade_c))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"D"'], fill=grade_d))
		ws.conditional_formatting.add('G2:G9999', CellIsRule(operator='equal', formula=['"F"'], fill=grade_f))

	if config.pad and config.keyboard:
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"SSS"'], fill=grade_sss))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"SS"'], fill=grade_ss))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"S"'], fill=grade_s))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"A"'], fill=grade_a))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"B"'], fill=grade_b))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"C"'], fill=grade_c))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"D"'], fill=grade_d))
		ws.conditional_formatting.add('K2:K9999', CellIsRule(operator='equal', formula=['"F"'], fill=grade_f))

	ws.freeze_panes = 'C2'

def write_summary_sheet(ws, db, chart_set, config, mixId, pad):
	table_names = ["Single + Single Performance", "Double + Double Performance", "Half-Double", "Routine", "Co-Op"]
	table_colors = ["ff2222", "22dd22", "cc0088", "44ccff", "ffdd00"]

	def difficulty_sort_key(d):
		if d == None:
			return 99
		if config.down:
			return -d
		return d

	def difficulty_name(d):
		if d == None:
			return "??"
		return "%02d" % d

	table = []
	for cid in chart_set:
		ver = db.chart_version_in_mix(cid, mixId)

		mode_title = db.chart_mode_str(cid, ver)
		if mode_title == None: continue
		mode_title = mode_title.lower()

		table_num = -1
		if mode_title.startswith("single"):
			table_num = 0
		elif mode_title.startswith("double"):
			table_num = 1
		elif mode_title.startswith("half"):
			table_num = 2
		elif mode_title.startswith("routine"):
			table_num = 3
		elif mode_title.startswith("co"):
			table_num = 4
		if table_num == -1: continue

		difficulty = db.chart_difficulty(cid, ver)

		table.append((table_num, difficulty_sort_key(difficulty), difficulty, cid))

	table.sort()
	table.append((None, None, None, None))

	side_left = {}
	side_down = {}
	side_right = {}
	side_up = {}

	ws.cell(row=1, column=17, value="CID")
	ws.cell(row=1, column=18, value="T")
	ws.cell(row=1, column=19, value="D")
	ws.cell(row=1, column=20, value="Pass")
	ws.cell(row=1, column=21, value="Grade")
	ws.cell(row=1, column=22, value="Miss")

	last_mode = last_diff = -1
	start_row = end_row = -1
	main_row = 1

	score_col = 6
	if not pad and config.pad:
		score_col = 10

	for r, (mode, _, diff, cid) in enumerate(table):
		# Add an entry to the hidden lookup table
		if mode != None:
			ws.cell(row=r+2, column=17, value=cid)
			ws.cell(row=r+2, column=18, value=mode)
			ws.cell(row=r+2, column=19, value=difficulty_name(diff))
			ws.cell(row=r+2, column=20, value='=IF(VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE)="", "", UPPER(VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE)))' % (r+2, score_col+0, r+2, score_col+0))
			ws.cell(row=r+2, column=21, value='=IF(VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE)="", "", UPPER(VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE)))' % (r+2, score_col+1, r+2, score_col+1))
			ws.cell(row=r+2, column=22, value='=IF(VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE)="", "", VLOOKUP(Q%d, Scores!$A$2:$M$9999, %d, FALSE))' % (r+2, score_col+2, r+2, score_col+2))

		# If there is a previous row...
		if last_mode != -1:
			# Finish writing the previous row
			if last_diff != diff or last_mode != mode:
				# TODO: Borders, Alignment, Formatting on Avg
				end_row = r+1
				ws.cell(row=main_row-1, column=2,  value='=COUNTIF(T%d:T%d, "Y")' % (start_row, end_row))
				ws.cell(row=main_row-1, column=3,  value='=COUNTIF(T%d:T%d, "N")' % (start_row, end_row))
				ws.cell(row=main_row-1, column=4,  value='=(%d-%d+1) - B%d - C%d' % (end_row, start_row, main_row-1, main_row-1))
				ws.cell(row=main_row-1, column=5,  value='=IF(COUNTIF(U%d:U%d, "SSS")<>0,COUNTIF(U%d:U%d, "SSS"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=6,  value='=IF(COUNTIF(U%d:U%d, "SS")<>0,COUNTIF(U%d:U%d, "SS"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=7,  value='=IF(COUNTIF(U%d:U%d, "S")<>0,COUNTIF(U%d:U%d, "S"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=8,  value='=IF(COUNTIF(U%d:U%d, "A")<>0,COUNTIF(U%d:U%d, "A"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=9,  value='=IF(COUNTIF(U%d:U%d, "B")<>0,COUNTIF(U%d:U%d, "B"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=10, value='=IF(COUNTIF(U%d:U%d, "C")<>0,COUNTIF(U%d:U%d, "C"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=11, value='=IF(COUNTIF(U%d:U%d, "D")<>0,COUNTIF(U%d:U%d, "D"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=12, value='=IF(COUNTIF(U%d:U%d, "F")<>0,COUNTIF(U%d:U%d, "F"),"")' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=13, value='=IF(SUMPRODUCT(--(V%d:V%d<>""))=0,"",MIN(V%d:V%d))' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=14, value='=IF(SUMPRODUCT(--(V%d:V%d<>""))=0,"",MAX(V%d:V%d))' % (start_row, end_row, start_row, end_row))
				ws.cell(row=main_row-1, column=15, value='=IF(SUMPRODUCT(--(V%d:V%d<>""))=0,"",AVERAGE(V%d:V%d))' % (start_row, end_row, start_row, end_row))

			# Finish the table
			if last_mode != mode:
				# TODO: Borders
				main_row += 1

		# If this is a new entry...
		if mode != None:
			# Draw the header of a new table
			if last_mode != mode:
				# TODO: Color, Merge, Center, Bold, Borders
				ws.cell(row=main_row,   column=1, value=table_names[mode])
				ws.cell(row=main_row+1, column=2, value="Passed") # Merge this into array thing
				ws.cell(row=main_row+1, column=3, value="Failed")
				ws.cell(row=main_row+1, column=4, value="Unplayed")
				ws.cell(row=main_row+1, column=5, value="SSS")
				ws.cell(row=main_row+1, column=6, value="SS")
				ws.cell(row=main_row+1, column=7, value="S")
				ws.cell(row=main_row+1, column=8, value="A")
				ws.cell(row=main_row+1, column=9, value="B")
				ws.cell(row=main_row+1, column=10, value="C")
				ws.cell(row=main_row+1, column=11, value="D")
				ws.cell(row=main_row+1, column=12, value="F")
				ws.cell(row=main_row+1, column=13, value="Low Miss")
				ws.cell(row=main_row+1, column=14, value="High Miss")
				ws.cell(row=main_row+1, column=15, value="Avg Miss")
				main_row += 2

			# Draw the next row
			if last_mode != mode or last_diff != diff:
				# TODO: Color, Borders
				ws.cell(row=main_row, column=1, value=difficulty_name(diff))
				start_row = r+2
				main_row += 1

		(last_mode, last_diff) = (mode, diff)

	# TODO: Set borders
	# TODO: Hide rows Q-V

	widths = [3, 6, 6, 8, 3, 3, 3, 3, 3, 3, 3, 3, 8, 9, 8, 8, 5, 2, 3, 4, 5, 4]
	for i, w in enumerate(widths):
		ws.column_dimensions[gcl(i+1)].width = w

def generate_xlsx(dbpath, xlpath, config):
	print("Reading database file...")
	db = read_database(dbpath)

	config.mix_ids = titles_to_ids(config.mixes, db.mixes, "mix")
	config.mode_ids = titles_to_ids(config.modes, db.modes, "mode")

	mix_to_charts = {}
	all_filtered_charts = set()
	for mid in config.mix_ids:
		charts = set()
		for cid in db.charts.keys():
			ver = db.chart_version_in_mix(cid, mid)
			if ver != None:
				rating = db.chart_rating(cid, ver)
				if rating.mode in config.mode_ids:
					diff = rating.difficulty
					if diff == None and config.unrated:
						charts.add(cid)
						continue
					if diff != None and (diff >= config.diff_min and diff <= config.diff_max):
						charts.add(cid)
						continue
		mix_to_charts[mid] = charts
		all_filtered_charts |= charts

	latest_filtered_mix = -1
	latest_mix = -1
	for mid in db.mixes:
		if latest_mix == -1:
			latest_mix = mid
		elif db.mixes[mid].order > db.mixes[latest_mix].order:
			latest_mix = mid
	for mid in config.mix_ids:
		if latest_filtered_mix == -1:
			latest_filtered_mix = mid
		elif db.mixes[mid].order > db.mixes[latest_filtered_mix].order:
			latest_filtered_mix = mid

	print("Creating score sheet...")
	wb = Workbook()
	ws_scores = wb.active
	ws_scores.title = "Scores"
	write_score_sheet(ws_scores, db, all_filtered_charts, config, latest_filtered_mix)

	if config.pad:
		for mid in config.mix_ids:
			print("Creating summary sheet (Pad, %s)..." % db.mixes[mid].title)
			ws_summary = wb.create_sheet(title=("Summary (Pad) %s" % db.mixes[mid].title)[:31])
			write_summary_sheet(ws_summary, db, mix_to_charts[mid], config, mid, True)
	if config.keyboard:
		for mid in config.mix_ids:
			print("Creating summary sheet (Kbd, %s)..." % db.mixes[mid].title)
			ws_summary = wb.create_sheet(title=("Summary (Kbd) %s" % db.mixes[mid].title)[:31])
			write_summary_sheet(ws_summary, db, mix_to_charts[mid], config, mid, False)

	print("Creating data sheet...")
	ws_dump = wb.create_sheet(title="Data")
	write_data_sheet(ws_dump, db, all_filtered_charts, config)

	print("Creating complete data sheet...")
	ws_dump = wb.create_sheet(title="Data (Complete)")
	write_data_sheet(ws_dump, db, set(db.charts), None)

	print("Creating about sheet...")
	options = []
	if config.pad: options += ["+Pad"]
	if config.keyboard: options += ["+Keyboard"]

	bold = Font(bold=True)
	ws_marker = wb.create_sheet(title="About")
	ws_marker.cell(row=1, column=1, value="Database Name:").font = bold
	ws_marker.cell(row=1, column=2, value=dbpath)
	ws_marker.cell(row=2, column=1, value="Latest Mix in Database:").font = bold
	ws_marker.cell(row=2, column=2, value=db.version_title(db.latest_version()))
	ws_marker.cell(row=3, column=1, value="Sheet Generated On:").font = bold
	ws_marker.cell(row=3, column=2, value=str(datetime.datetime.now()))
	ws_marker.cell(row=4, column=1, value="Generator Version:").font = bold
	ws_marker.cell(row=4, column=2, value="v0.5")

	ws_marker.cell(row=6, column=1, value="Player:").font = bold
	ws_marker.cell(row=6, column=2, value="[YOUR NAME HERE]").font = bold
	ws_marker.cell(row=7, column=1, value="Mixes:").font = bold
	ws_marker.cell(row=7, column=2, value=", ".join(config.mixes))
	ws_marker.cell(row=8, column=1, value="Modes:").font = bold
	ws_marker.cell(row=8, column=2, value=", ".join(config.modes))
	ws_marker.cell(row=9, column=1, value="Difficulties:").font = bold
	ws_marker.cell(row=9, column=2, value="%d-%d%s" % (config.diff_min, config.diff_max, (""," (+Unrated)")[config.unrated]))
	ws_marker.cell(row=10, column=1, value="Options:").font = bold
	ws_marker.cell(row=10, column=2, value="%s" % ", ".join(options))
	
	adjust_column_widths(ws_marker, range(1,2+1), range(1,10+1))

	print("Saving workbook...")
	wb.save(xlpath)
	wb.close()

if __name__ == '__main__':
	print("Reading config file...")
	config = parse_config("config.txt")
	generate_xlsx(DBPATH, "output.xlsx", config)
